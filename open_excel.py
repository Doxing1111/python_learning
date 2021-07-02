'''
date: 2021/07/02
author: Joe
'''

import openpyxl

# 建立一個工作簿
workbook = openpyxl.Workbook()

# 取得第 1 個工作表
sheet = workbook.worksheets[0]

# 以儲存格位置寫入資料
sheet['A1'] = '一年甲班'

# 以串列寫入資料
listtitle = ['座號', '姓名', '國文', '英文', '數學']
sheet.append(listtitle)
listdatas = [[1, 'Joe', 100, 90, 80],
    [2, 'Mark', 80, 75, 60],
    [3, 'Allen', 60, 55, 40]]

for listdata in listdatas:
    sheet.append(listdata)

# 儲存檔案
workbook.save('test.xlsx')

# ==============================================================
#  讀取檔案
workbook = openpyxl.load_workbook('test.xlsx')

# 取得第 1 個工作表
sheet = workbook.worksheets[0]

# 取得指定儲存格
print(sheet['A1'], sheet['A1'].value)

# 取得總行、列數
print(sheet.max_row, sheet.max_column)

# 顯示 cell資料
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value,end="   ")
    print()

sheet['A1'] = '二年甲班'
workbook.save('test.xlsx')

